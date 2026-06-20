import csv
import json

from openpyxl import load_workbook

from app.exporter import (
    CATEGORY_FIELDS,
    FAILED_URL_FIELDS,
    OUTPUT_FIELDS,
    export_records,
    export_trade_categories,
)
from app.models import BusinessRecord, FailedURL, TradeCategory


def test_export_records_uses_required_columns_and_empty_failed_url_headers(tmp_path):
    record = BusinessRecord(
        business_name="Example Builder",
        trade_category="Building",
        region="Auckland",
        phone="+64 9 555 1111",
        description="Residential building services.",
        listing_url="https://directory.example/example-builder",
        source_url="https://directory.example/",
    )
    csv_path = tmp_path / "records.csv"
    excel_path = tmp_path / "records.xlsx"
    summary_path = tmp_path / "summary.json"
    failed_path = tmp_path / "failed_urls.csv"
    categories_path = tmp_path / "trade_categories.csv"
    export_trade_categories(
        [
            TradeCategory(
                trade_category="Building",
                category_url="https://directory.example/category/building",
                listed_count=1,
            )
        ],
        output_path=categories_path,
    )

    export_records(
        [record],
        {"final_records": 1},
        excel_path=excel_path,
        csv_path=csv_path,
        summary_path=summary_path,
        failed_urls_path=failed_path,
        categories_path=categories_path,
    )

    with csv_path.open(encoding="utf-8", newline="") as file:
        rows = list(csv.DictReader(file))
    assert rows[0]["phone"] == "+64 9 555 1111"
    assert isinstance(rows[0]["phone"], str)
    assert "e+" not in rows[0]["phone"].lower()

    raw_csv = csv_path.read_text(encoding="utf-8")
    assert '"+64 9 555 1111"' in raw_csv

    workbook = load_workbook(excel_path)
    assert workbook.sheetnames == [
        "Master",
        "Summary",
        "Failed URLs",
        "Categories",
        "Data Quality Notes",
    ]
    assert [cell.value for cell in workbook["Master"][1]] == OUTPUT_FIELDS
    phone_cell = workbook["Master"]["D2"]
    assert phone_cell.value == "+64 9 555 1111"
    assert isinstance(phone_cell.value, str)
    assert phone_cell.data_type == "s"
    assert phone_cell.number_format == "@"
    assert "e+" not in phone_cell.value.lower()

    assert workbook["Master"].freeze_panes == "A2"
    assert workbook["Master"].auto_filter.ref == "A1:G2"
    assert workbook["Master"]["E2"].alignment.wrap_text is True
    assert workbook["Master"]["A1"].font.bold is True
    assert workbook["Failed URLs"]["A2"].value == (
        "No failed URLs recorded for this sample run."
    )
    assert workbook["Categories"]["A2"].value == "Building"
    assert workbook["Data Quality Notes"]["B11"].value == "PASS"

    with failed_path.open(encoding="utf-8", newline="") as file:
        assert next(csv.reader(file)) == FAILED_URL_FIELDS

    assert json.loads(summary_path.read_text(encoding="utf-8"))["final_records"] == 1


def test_export_records_writes_failed_urls(tmp_path):
    failed_url = FailedURL(
        url="https://directory.example/broken",
        error_type="HTTPError",
        error_message="503 Server Error",
        status_code=503,
        timestamp="2026-01-01T00:00:00+00:00",
    )
    failed_path = tmp_path / "failed_urls.csv"

    export_records(
        [],
        {},
        failed_urls=[failed_url],
        excel_path=tmp_path / "records.xlsx",
        csv_path=tmp_path / "records.csv",
        summary_path=tmp_path / "summary.json",
        failed_urls_path=failed_path,
        categories_path=tmp_path / "missing_categories.csv",
    )

    with failed_path.open(encoding="utf-8", newline="") as file:
        rows = list(csv.DictReader(file))

    assert rows[0]["url"] == failed_url.url
    assert rows[0]["status_code"] == "503"

    workbook = load_workbook(tmp_path / "records.xlsx")
    assert workbook["Failed URLs"]["A2"].value == failed_url.url
    assert workbook["Categories"]["A2"].value == (
        "Run discover-categories mode to populate this sheet."
    )
    assert workbook["Summary"]["A2"].value == "Note"
    assert "No scraping_summary.json data" in workbook["Summary"]["B2"].value


def test_export_trade_categories_uses_discovery_schema(tmp_path):
    output_path = tmp_path / "trade_categories.csv"
    export_trade_categories(
        [
            TradeCategory(
                trade_category="Electricians",
                category_url=(
                    "https://tradehq.co.nz/business-category/electricians/"
                ),
                listed_count=89,
            )
        ],
        output_path=output_path,
    )

    with output_path.open(encoding="utf-8", newline="") as file:
        rows = list(csv.DictReader(file))

    assert list(rows[0].keys()) == CATEGORY_FIELDS
    assert rows[0]["trade_category"] == "Electricians"
    assert rows[0]["listed_count"] == "89"
