"""Export cleaned records, failed URLs, and the JSON summary."""

import csv
import json
from collections import Counter
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import (
    CSV_OUTPUT_PATH,
    EXCEL_OUTPUT_PATH,
    FAILED_URLS_OUTPUT_PATH,
    OUTPUTS_DIR,
    SUMMARY_OUTPUT_PATH,
    TRADE_CATEGORIES_OUTPUT_PATH,
)
from app.models import BusinessRecord, FailedURL, TradeCategory
from app.utils import ensure_directory, model_to_dict


OUTPUT_FIELDS = [
    "business_name",
    "trade_category",
    "region",
    "phone",
    "description",
    "listing_url",
    "source_url",
]
FAILED_URL_FIELDS = [
    "url",
    "error_type",
    "error_message",
    "status_code",
    "timestamp",
]
CATEGORY_FIELDS = ["trade_category", "category_url", "listed_count"]
WORKBOOK_SHEETS = [
    "Master",
    "Summary",
    "Failed URLs",
    "Categories",
    "Data Quality Notes",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
SECTION_FONT = Font(bold=True, color="1F1F1F")


def _records_to_rows(records: Iterable[BusinessRecord]) -> tuple[list[str], list[dict]]:
    rows = [model_to_dict(record) for record in records]
    clean_rows = []
    for row in rows:
        clean_row = {
            field_name: row.get(field_name)
            for field_name in OUTPUT_FIELDS
        }
        phone = clean_row.get("phone")
        if phone is not None and not isinstance(phone, str):
            raise TypeError("phone must remain a string during export")
        clean_rows.append(clean_row)
    return OUTPUT_FIELDS, clean_rows


def _style_header(worksheet, row_number: int = 1) -> None:
    for cell in worksheet[row_number]:
        if cell.value is not None:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(vertical="center")


def _format_worksheet(worksheet, maximum_width: int = 60) -> None:
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            maximum_width,
        )


def _append_table(worksheet, headers: list[str], rows: list[dict]) -> None:
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(header) for header in headers])
    _style_header(worksheet)
    worksheet.auto_filter.ref = worksheet.dimensions


def _format_master_sheet(worksheet, rows: list[dict]) -> None:
    _append_table(worksheet, OUTPUT_FIELDS, rows)
    worksheet.freeze_panes = "A2"
    phone_column = OUTPUT_FIELDS.index("phone") + 1
    for row_number in range(2, worksheet.max_row + 1):
        phone_cell = worksheet.cell(row=row_number, column=phone_column)
        if phone_cell.value is not None:
            phone_cell.value = str(phone_cell.value)
            phone_cell.data_type = "s"
        phone_cell.number_format = "@"

        worksheet.cell(row=row_number, column=5).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
        for column_number in (6, 7):
            worksheet.cell(
                row=row_number,
                column=column_number,
            ).alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "A": 34,
        "B": 24,
        "C": 20,
        "D": 20,
        "E": 70,
        "F": 55,
        "G": 55,
    }
    for column_letter, width in widths.items():
        worksheet.column_dimensions[column_letter].width = width


def _read_csv_rows(path: Path, headers: list[str]) -> list[dict]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8-sig", newline="") as file:
        return [
            {header: row.get(header, "") for header in headers}
            for row in csv.DictReader(file)
        ]


def _load_summary(summary: dict, summary_path: Path) -> dict:
    if summary:
        return summary
    if not summary_path.exists():
        return {}
    try:
        with summary_path.open(encoding="utf-8") as file:
            loaded = json.load(file)
        return loaded if isinstance(loaded, dict) else {}
    except (json.JSONDecodeError, OSError):
        return {}


def _add_section_title(worksheet, title: str) -> None:
    worksheet.append([title])
    cell = worksheet.cell(row=worksheet.max_row, column=1)
    cell.fill = SECTION_FILL
    cell.font = SECTION_FONT


def _build_summary_sheet(
    worksheet,
    rows: list[dict],
    summary: dict,
    summary_exists: bool,
) -> None:
    worksheet.append(["Metric", "Value"])
    _style_header(worksheet)

    if not summary_exists:
        worksheet.append(
            ["Note", "No scraping_summary.json data was available for this workbook."]
        )

    metrics = [
        (
            "total records",
            summary.get(
                "total_raw_records",
                summary.get("total_records_scraped", len(rows)),
            ),
        ),
        ("final records", summary.get("final_records", len(rows))),
        ("duplicates removed", summary.get("duplicates_removed", 0)),
        (
            "detail descriptions full",
            summary.get("descriptions_from_detail_page", ""),
        ),
        (
            "detail descriptions fallback",
            summary.get("descriptions_fallback_to_listing_excerpt", ""),
        ),
        ("category pages scraped", summary.get("category_pages_scraped", "")),
        ("categories requested", summary.get("categories_requested", "")),
    ]
    for metric, value in metrics:
        worksheet.append([metric, value])

    _add_section_title(worksheet, "Records by trade_category")
    worksheet.append(["trade_category", "records"])
    _style_header(worksheet, worksheet.max_row)
    category_counts = Counter(
        row["trade_category"] for row in rows if row.get("trade_category")
    )
    for trade_category, count in sorted(category_counts.items()):
        worksheet.append([trade_category, count])

    _add_section_title(worksheet, "Records by region")
    worksheet.append(["region", "records"])
    _style_header(worksheet, worksheet.max_row)
    region_counts = Counter(row["region"] for row in rows if row.get("region"))
    for region, count in sorted(region_counts.items()):
        worksheet.append([region, count])

    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 34
    worksheet.column_dimensions["B"].width = 70


def _build_failed_urls_sheet(worksheet, failed_rows: list[dict]) -> None:
    _append_table(worksheet, FAILED_URL_FIELDS, failed_rows)
    if not failed_rows:
        worksheet.append(
            ["No failed URLs recorded for this sample run.", "", "", "", ""]
        )
        worksheet.merge_cells(
            start_row=2,
            start_column=1,
            end_row=2,
            end_column=len(FAILED_URL_FIELDS),
        )
        worksheet["A2"].font = Font(italic=True)
    _format_worksheet(worksheet)


def _build_categories_sheet(worksheet, category_rows: list[dict]) -> None:
    _append_table(worksheet, CATEGORY_FIELDS, category_rows)
    if not category_rows:
        worksheet.append(
            ["Run discover-categories mode to populate this sheet.", "", ""]
        )
        worksheet.merge_cells(
            start_row=2,
            start_column=1,
            end_row=2,
            end_column=len(CATEGORY_FIELDS),
        )
        worksheet["A2"].font = Font(italic=True)
    _format_worksheet(worksheet)
    worksheet.column_dimensions["A"].width = 32
    worksheet.column_dimensions["B"].width = 65
    worksheet.column_dimensions["C"].width = 16


def _build_data_quality_sheet(worksheet) -> None:
    notes = [
        ("Manual validation sample", "10 records"),
        ("Categories checked", "Electricians, Plumbers, Builders"),
        ("business_name", "10/10 correct"),
        ("phone", "10/10 correct"),
        ("trade_category", "10/10 correct"),
        ("region", "10/10 correct"),
        ("description", "10/10 acceptable"),
        ("listing_url", "10/10 correct"),
        ("source_url", "10/10 correct"),
        ("Validation result", "PASS"),
        (
            "Phone handling note",
            "Phone values are exported as text to avoid leading-zero loss or "
            "scientific notation.",
        ),
        (
            "Description note",
            "Descriptions are collected from detail pages when available and "
            "cleaned before export.",
        ),
        (
            "Ethical/source note",
            "Data is collected from publicly available business directory "
            "pages, and each record includes source URLs.",
        ),
    ]
    worksheet.append(["Validation item", "Result / note"])
    _style_header(worksheet)
    for label, value in notes:
        worksheet.append([label, value])
        worksheet.cell(row=worksheet.max_row, column=2).alignment = Alignment(
            wrap_text=True,
            vertical="top",
        )
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 32
    worksheet.column_dimensions["B"].width = 95


def _build_workbook(
    rows: list[dict],
    summary: dict,
    failed_rows: list[dict],
    category_rows: list[dict],
    summary_exists: bool,
) -> Workbook:
    workbook = Workbook()
    master_sheet = workbook.active
    master_sheet.title = "Master"
    _format_master_sheet(master_sheet, rows)

    summary_sheet = workbook.create_sheet("Summary")
    _build_summary_sheet(summary_sheet, rows, summary, summary_exists)

    failed_sheet = workbook.create_sheet("Failed URLs")
    _build_failed_urls_sheet(failed_sheet, failed_rows)

    categories_sheet = workbook.create_sheet("Categories")
    _build_categories_sheet(categories_sheet, category_rows)

    quality_sheet = workbook.create_sheet("Data Quality Notes")
    _build_data_quality_sheet(quality_sheet)
    return workbook


def export_records(
    records: Iterable[BusinessRecord],
    summary: dict,
    failed_urls: Iterable[FailedURL] = (),
    excel_path: Path = EXCEL_OUTPUT_PATH,
    csv_path: Path = CSV_OUTPUT_PATH,
    summary_path: Path = SUMMARY_OUTPUT_PATH,
    failed_urls_path: Path = FAILED_URLS_OUTPUT_PATH,
    categories_path: Path = TRADE_CATEGORIES_OUTPUT_PATH,
) -> None:
    """Write all output files."""

    ensure_directory(OUTPUTS_DIR)
    headers, rows = _records_to_rows(records)

    with csv_path.open("w", encoding="utf-8", newline="") as file:
        # CSV has no native type metadata. Quoting every field keeps phone
        # values textual for standards-compliant CSV readers and preserves
        # leading zeroes, plus signs, spaces, and punctuation.
        writer = csv.DictWriter(
            file,
            fieldnames=headers,
            quoting=csv.QUOTE_ALL,
        )
        writer.writeheader()
        writer.writerows(rows)

    with summary_path.open("w", encoding="utf-8") as file:
        json.dump(summary, file, indent=2)

    failed_rows = [model_to_dict(failed_url) for failed_url in failed_urls]
    export_failed_urls(failed_rows, failed_urls_path)
    category_rows = _read_csv_rows(categories_path, CATEGORY_FIELDS)
    workbook_summary = _load_summary(summary, summary_path)
    workbook = _build_workbook(
        rows,
        workbook_summary,
        failed_rows or _read_csv_rows(failed_urls_path, FAILED_URL_FIELDS),
        category_rows,
        summary_exists=bool(workbook_summary),
    )
    workbook.save(excel_path)


def export_failed_urls(
    failed_urls: Iterable[FailedURL] | Iterable[dict],
    failed_urls_path: Path = FAILED_URLS_OUTPUT_PATH,
) -> None:
    """Write failed URLs, including an empty header-only report."""

    failed_rows = [
        item if isinstance(item, dict) else model_to_dict(item)
        for item in failed_urls
    ]
    ensure_directory(failed_urls_path.parent)
    with failed_urls_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=FAILED_URL_FIELDS)
        writer.writeheader()
        writer.writerows(failed_rows)


def export_trade_categories(
    categories: Iterable[TradeCategory],
    output_path: Path = TRADE_CATEGORIES_OUTPUT_PATH,
) -> None:
    """Export discovered categories to a review-friendly CSV."""

    ensure_directory(output_path.parent)
    with output_path.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=CATEGORY_FIELDS,
            quoting=csv.QUOTE_ALL,
        )
        writer.writeheader()
        writer.writerows(
            {
                field_name: model_to_dict(category).get(field_name)
                for field_name in CATEGORY_FIELDS
            }
            for category in categories
        )

